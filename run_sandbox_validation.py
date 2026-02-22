#!/usr/bin/env python3
"""
One-command sandbox run for the RMBS model.

This runner:
1) creates a timestamped sandbox directory under .sandbox_runs
2) runs workbook build + v1/v2 engines into sandbox outputs
3) optionally automates Excel scenario exports for CompareExport (macOS Excel)
4) runs 3-scenario Excel/Python reconciliation when exports are available
5) compiles LaTeX docs into sandbox outputs
6) computes an excellence scorecard from sandbox artifacts only

Key optional controls:
- --scenario-config <yaml>: load calibrated v1/v2 scenarios
- --output-root <path>: write timestamped runs under a custom subfolder
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import platform
import shutil
import subprocess
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable


SCENARIOS = ["Base", "Mild_Stress", "Severe_Stress"]


@dataclass
class CmdResult:
    name: str
    returncode: int
    log_path: Path


def run_cmd(
    cmd: list[str],
    *,
    name: str,
    log_path: Path,
    cwd: Path | None = None,
    env: dict[str, str] | None = None,
) -> CmdResult:
    log_path.parent.mkdir(parents=True, exist_ok=True)
    with log_path.open("w", encoding="utf-8") as log:
        log.write(f"$ {' '.join(cmd)}\n")
        proc = subprocess.run(
            cmd,
            cwd=str(cwd) if cwd else None,
            env=env,
            stdout=log,
            stderr=subprocess.STDOUT,
            check=False,
            text=True,
        )
        log.write(f"\n[exit_code] {proc.returncode}\n")
    return CmdResult(name=name, returncode=proc.returncode, log_path=log_path)


def ensure_files(paths: Iterable[Path]) -> None:
    missing = [str(p) for p in paths if not p.exists()]
    if missing:
        raise FileNotFoundError(f"Missing required files:\n- " + "\n- ".join(missing))


def write_csv(path: Path, header: list[str], rows: list[list[object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(header)
        writer.writerows(rows)


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def export_compare_csv_from_workbook(workbook_path: Path, out_csv: Path) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    ws = wb["CompareExport"]
    out_csv.parent.mkdir(parents=True, exist_ok=True)
    with out_csv.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        for r in range(1, 362):
            writer.writerow([ws.cell(row=r, column=c).value for c in range(1, 14)])


def automate_excel_scenario_recalc(workbook_path: Path, scenario: str) -> None:
    if platform.system() != "Darwin":
        raise RuntimeError("Excel automation is currently implemented only for macOS.")
    if shutil.which("osascript") is None:
        raise RuntimeError("osascript not found; cannot automate Excel.")

    script = f"""
tell application "Microsoft Excel"
\tset wb to open workbook workbook file name (POSIX file "{workbook_path}")
\tset value of range "B12" of worksheet "Inputs" of wb to "{scenario}"
\tcalculate full
\tclose wb saving yes
end tell
"""
    subprocess.run(["osascript", "-e", script], check=True, text=True)


def parse_month_coverage_from_report(path: Path) -> str:
    if not path.exists():
        return "UNKNOWN"
    for line in path.read_text(encoding="utf-8").splitlines():
        if "Month coverage status:" in line:
            return line.split(":", 1)[1].strip()
    return "UNKNOWN"


def float_or_zero(value: str) -> float:
    try:
        return float(value)
    except Exception:
        return 0.0


def build_scorecard(sb: Path, strict: float, warn: float) -> dict:
    def weighted_diff_metric(weight: float, values: list[float]) -> dict:
        maxv = max(abs(v) for v in values)
        if maxv <= strict:
            return {
                "weight": weight,
                "score": weight,
                "status": "PASS",
                "detail": f"max={maxv:.12g} <= {strict}",
            }
        if maxv <= warn:
            return {
                "weight": weight,
                "score": weight / 2.0,
                "status": "WARN",
                "detail": f"max={maxv:.12g} <= {warn}",
            }
        return {
            "weight": weight,
            "score": 0.0,
            "status": "FAIL",
            "detail": f"max={maxv:.12g} > {warn}",
        }

    def weighted_floor_metric(weight: float, value: float) -> dict:
        if value >= -strict:
            return {
                "weight": weight,
                "score": weight,
                "status": "PASS",
                "detail": f"min={value:.12g} >= {-strict}",
            }
        if value >= -warn:
            return {
                "weight": weight,
                "score": weight / 2.0,
                "status": "WARN",
                "detail": f"min={value:.12g} >= {-warn}",
            }
        return {
            "weight": weight,
            "score": 0.0,
            "status": "FAIL",
            "detail": f"min={value:.12g} < {-warn}",
        }

    v1_summary = read_csv(sb / "outputs/v1/rmbs_validation_summary.csv")
    v2_summary = read_csv(sb / "outputs/v2/rmbs_v2_validation_summary.csv")
    v1_by = {r["scenario"]: r for r in v1_summary}
    v2_by = {r["scenario"]: r for r in v2_summary}

    metrics: dict[str, dict] = {}

    vals: list[float] = []
    for s in SCENARIOS:
        vals.extend(
            [
                float_or_zero(v1_by[s]["principal_diff_max"]),
                float_or_zero(v1_by[s]["interest_diff_max"]),
                float_or_zero(v1_by[s]["loss_diff_max"]),
                float_or_zero(v2_by[s]["principal_diff_max"]),
                float_or_zero(v2_by[s]["interest_diff_max"]),
                float_or_zero(v2_by[s]["loss_diff_max"]),
            ]
        )
    metrics["E1_core_closure"] = weighted_diff_metric(15, vals)
    metrics["E2_pool_roll"] = weighted_diff_metric(
        5, [float_or_zero(v2_by[s]["pool_roll_diff_max"]) for s in SCENARIOS]
    )
    min_bal = min(
        min(float_or_zero(v1_by[s]["min_balance_seen"]) for s in SCENARIOS),
        min(float_or_zero(v2_by[s]["min_balance_seen"]) for s in SCENARIOS),
    )
    metrics["E3_non_negative"] = weighted_floor_metric(5, min_bal)
    metrics["E4_unallocated_buckets"] = weighted_diff_metric(
        5,
        [float_or_zero(v2_by[s]["max_unallocated_principal"]) for s in SCENARIOS]
        + [float_or_zero(v2_by[s]["max_unallocated_loss"]) for s in SCENARIOS],
    )

    compare_status_rows = read_csv(sb / "logs/compare_status.csv")
    status_by = {r["scenario"]: r for r in compare_status_rows}
    all_compare_ran = all(status_by.get(s, {}).get("status") == "RAN" for s in SCENARIOS)

    compare_details: dict[str, dict] = {}
    for s in SCENARIOS:
        item = {"status": status_by.get(s, {}).get("status", "MISSING_STATUS")}
        if item["status"] == "RAN":
            diff_path = sb / f"outputs/compare/excel_python_{s}_diffs.csv"
            report_path = sb / f"outputs/compare/excel_python_{s}_report.md"
            diffs = read_csv(diff_path)
            max_abs = max((float_or_zero(r["abs_diff"]) for r in diffs), default=0.0)
            fail_rows = sum(1 for r in diffs if r["status"] != "PASS")
            cov = parse_month_coverage_from_report(report_path)
            item.update({"max_abs_diff": max_abs, "fail_rows": fail_rows, "coverage": cov})
        compare_details[s] = item

    if all_compare_ran:
        metrics["E5_compare_max_diff"] = weighted_diff_metric(
            18, [compare_details[s]["max_abs_diff"] for s in SCENARIOS]
        )
    else:
        metrics["E5_compare_max_diff"] = {
            "weight": 18,
            "score": 0.0,
            "status": "FAIL",
            "detail": "missing one or more manual/automated Excel exports",
        }

    if all_compare_ran and all(compare_details[s]["fail_rows"] == 0 for s in SCENARIOS):
        metrics["E6_compare_fail_rows"] = {
            "weight": 9,
            "score": 9.0,
            "status": "PASS",
            "detail": "all scenarios have 0 FAIL rows",
        }
    else:
        metrics["E6_compare_fail_rows"] = {
            "weight": 9,
            "score": 0.0,
            "status": "FAIL",
            "detail": "compare not complete or FAIL rows present",
        }

    if all_compare_ran and all(str(compare_details[s]["coverage"]).upper() == "PASS" for s in SCENARIOS):
        metrics["E7_compare_coverage"] = {
            "weight": 6,
            "score": 6.0,
            "status": "PASS",
            "detail": "coverage PASS for all scenarios",
        }
    else:
        metrics["E7_compare_coverage"] = {
            "weight": 6,
            "score": 0.0,
            "status": "FAIL",
            "detail": "compare coverage incomplete",
        }

    v1_ok = (
        float_or_zero(v1_by["Severe_Stress"]["cum_loss"])
        >= float_or_zero(v1_by["Mild_Stress"]["cum_loss"])
        >= float_or_zero(v1_by["Base"]["cum_loss"])
    )
    v2_ok = (
        float_or_zero(v2_by["Severe_Stress"]["cum_loss"])
        >= float_or_zero(v2_by["Mild_Stress"]["cum_loss"])
        >= float_or_zero(v2_by["Base"]["cum_loss"])
        and float_or_zero(v2_by["Severe_Stress"]["max_dq_ratio"])
        >= float_or_zero(v2_by["Mild_Stress"]["max_dq_ratio"])
        >= float_or_zero(v2_by["Base"]["max_dq_ratio"])
    )
    metrics["E8_stress_monotonicity"] = {
        "weight": 10,
        "score": 10.0 if (v1_ok and v2_ok) else 0.0,
        "status": "PASS" if (v1_ok and v2_ok) else "FAIL",
        "detail": f"v1={v1_ok}, v2={v2_ok}",
    }

    v2_monthly = read_csv(sb / "outputs/v2/rmbs_v2_validation_monthly.csv")
    monthly_by = {s: [r for r in v2_monthly if r["scenario"] == s] for s in SCENARIOS}
    for s in SCENARIOS:
        monthly_by[s].sort(key=lambda r: int(float(r["month"])))

    threshold_loss = 0.03
    threshold_dq = 0.06

    def first_trigger_month(rows: list[dict[str, str]]) -> int:
        for r in rows:
            if int(float(r["trigger_on"])) == 1:
                return int(float(r["month"]))
        return -1

    def first_breach_month(rows: list[dict[str, str]]) -> int:
        if not rows:
            return -1
        pool0 = float_or_zero(rows[0]["pool_begin"])
        cum_loss = 0.0
        for r in rows:
            cum_loss += float_or_zero(r["loss"])
            dq = float_or_zero(r["dq_ratio"])
            if pool0 > 0 and ((cum_loss / pool0) >= threshold_loss or dq >= threshold_dq):
                return int(float(r["month"]))
        return -1

    trig = {s: first_trigger_month(monthly_by[s]) for s in SCENARIOS}
    breach = {s: first_breach_month(monthly_by[s]) for s in SCENARIOS}
    # Trigger correctness is evaluated as state coherence:
    # if a breach happens, trigger must turn on that month; if no breach, trigger must stay off.
    trigger_ok = all(trig[s] == breach[s] for s in SCENARIOS)
    metrics["E9_trigger_correctness"] = {
        "weight": 10,
        "score": 10.0 if trigger_ok else 0.0,
        "status": "PASS" if trigger_ok else "FAIL",
        "detail": f"trigger={trig}, breach={breach}",
    }

    core_logs = [
        sb / "logs/preflight.log",
        sb / "logs/build_workbook.log",
        sb / "logs/v1.log",
        sb / "logs/v2.log",
        sb / "logs/latex_status.csv",
    ]
    core_ok = all(p.exists() for p in core_logs)
    if core_ok and all_compare_ran:
        metrics["E10_reproducibility"] = {
            "weight": 7,
            "score": 7.0,
            "status": "PASS",
            "detail": "all logs present and compare complete",
        }
    elif core_ok:
        metrics["E10_reproducibility"] = {
            "weight": 7,
            "score": 3.5,
            "status": "WARN",
            "detail": "core logs present; compare pending",
        }
    else:
        metrics["E10_reproducibility"] = {
            "weight": 7,
            "score": 0.0,
            "status": "FAIL",
            "detail": "missing core logs",
        }

    v1_report = (sb / "outputs/v1/rmbs_validation_report.md").read_text(encoding="utf-8")
    v2_report = (sb / "outputs/v2/rmbs_v2_validation_report.md").read_text(encoding="utf-8")
    req_v1 = [
        "# RMBS Validation Report",
        "## Run Configuration",
        "## Scenario Summary",
        "## Integrity Checks",
        "## Interpretation",
    ]
    req_v2 = [
        "# RMBS v2 Validation Report",
        "## Configuration",
        "## Scenario Summary",
        "## Integrity Checks",
        "## Interpretation",
    ]
    has_v1 = all(x in v1_report for x in req_v1)
    has_v2 = all(x in v2_report for x in req_v2)
    if has_v1 and has_v2 and all_compare_ran:
        metrics["E11_governance"] = {
            "weight": 10,
            "score": 10.0,
            "status": "PASS",
            "detail": "v1/v2/compare governance artifacts complete",
        }
    elif has_v1 and has_v2:
        metrics["E11_governance"] = {
            "weight": 10,
            "score": 5.0,
            "status": "WARN",
            "detail": "v1/v2 governance complete; compare pending",
        }
    else:
        metrics["E11_governance"] = {
            "weight": 10,
            "score": 0.0,
            "status": "FAIL",
            "detail": "missing required report sections",
        }

    raw_total = sum(m["score"] for m in metrics.values())
    blocker = any(
        metrics[k]["status"] == "FAIL"
        for k in ["E1_core_closure", "E2_pool_roll", "E3_non_negative", "E4_unallocated_buckets"]
    )
    cap_active = not all_compare_ran
    cap_value = 89.0 if cap_active else None
    final_total = min(raw_total, cap_value) if cap_active else raw_total

    if blocker:
        grade = "BLOCKER"
    elif final_total >= 95 and not cap_active:
        grade = "Excellence"
    elif final_total >= 90:
        grade = "Strong"
    elif final_total >= 80:
        grade = "Acceptable"
    else:
        grade = "Needs remediation"

    shortfalls = []
    for key, metric in metrics.items():
        shortfalls.append((metric["weight"] - metric["score"], key, metric["detail"]))
    shortfalls.sort(reverse=True)

    return {
        "sandbox_path": str(sb),
        "strict_tolerance": strict,
        "warning_tolerance": warn,
        "metrics": metrics,
        "compare_status_by_scenario": compare_details,
        "raw_total": raw_total,
        "final_total": final_total,
        "grade": grade,
        "blocker": blocker,
        "cap_active": cap_active,
        "cap_value": cap_value,
        "top_gaps": [
            {"metric": k, "shortfall": s, "detail": d}
            for s, k, d in shortfalls
            if s > 0
        ][:5],
    }


def write_scorecard_files(sb: Path, scorecard: dict) -> tuple[Path, Path]:
    json_path = sb / "logs/excellence_scorecard.json"
    md_path = sb / "logs/excellence_scorecard.md"
    json_path.write_text(json.dumps(scorecard, indent=2), encoding="utf-8")

    lines: list[str] = []
    lines.append("# Sandbox Excellence Scorecard")
    lines.append(f"- Sandbox: `{scorecard['sandbox_path']}`")
    lines.append(f"- Raw score: {scorecard['raw_total']:.2f} / 100")
    if scorecard["cap_active"]:
        lines.append(
            f"- Capped score: {scorecard['final_total']:.2f} / 100 "
            f"(cap {scorecard['cap_value']:.0f} due to incomplete compare scenarios)"
        )
    else:
        lines.append(f"- Final score: {scorecard['final_total']:.2f} / 100")
    lines.append(f"- Grade: {scorecard['grade']}")
    lines.append(f"- Core blocker: {'YES' if scorecard['blocker'] else 'NO'}")
    lines.append("")
    lines.append("## Metric Results")
    for key, metric in scorecard["metrics"].items():
        lines.append(
            f"- {key}: {metric['status']} ({metric['score']:.2f}/{metric['weight']}) - {metric['detail']}"
        )
    lines.append("")
    lines.append("## Compare Status")
    for s in SCENARIOS:
        st = scorecard["compare_status_by_scenario"].get(s, {}).get("status", "UNKNOWN")
        lines.append(f"- {s}: {st}")
    lines.append("")
    lines.append("## Top Gaps")
    for gap in scorecard["top_gaps"]:
        lines.append(
            f"- {gap['metric']}: shortfall {gap['shortfall']:.2f} ({gap['detail']})"
        )

    md_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return json_path, md_path


def main() -> int:
    parser = argparse.ArgumentParser(description="Run full sandbox validation and scorecard.")
    parser.add_argument(
        "--root",
        default=str(Path(__file__).resolve().parent),
        help="Project root path (default: script directory).",
    )
    parser.add_argument("--run-id", default="", help="Optional run id. Default: timestamp.")
    parser.add_argument(
        "--output-root",
        default="",
        help="Optional output root for timestamped runs. Default: <root>/.sandbox_runs",
    )
    parser.add_argument(
        "--scenario-config",
        default="",
        help="Optional YAML path with v1_scenarios/v2_scenarios.",
    )
    parser.add_argument(
        "--automate-excel",
        action="store_true",
        help="Automate Excel scenario recalc/export via macOS AppleScript.",
    )
    parser.add_argument(
        "--skip-latex",
        action="store_true",
        help="Skip PDF compilation stage.",
    )
    parser.add_argument("--strict-tol", type=float, default=1e-6)
    parser.add_argument("--warn-tol", type=float, default=1e-4)
    args = parser.parse_args()

    root = Path(args.root).resolve()
    output_root = (
        Path(args.output_root).expanduser().resolve()
        if args.output_root
        else root / ".sandbox_runs"
    )
    scenario_config_path = (
        Path(args.scenario_config).expanduser().resolve() if args.scenario_config else None
    )
    if scenario_config_path and not scenario_config_path.exists():
        raise FileNotFoundError(f"Scenario config not found: {scenario_config_path}")

    run_id = args.run_id or datetime.now().strftime("%Y%m%d_%H%M%S")
    sb = output_root / run_id
    paths = {
        "inputs_exports": sb / "inputs/excel_exports",
        "out_v1": sb / "outputs/v1",
        "out_v2": sb / "outputs/v2",
        "out_compare": sb / "outputs/compare",
        "out_docs": sb / "outputs/docs",
        "logs": sb / "logs",
    }
    for p in paths.values():
        p.mkdir(parents=True, exist_ok=True)
    output_root.mkdir(parents=True, exist_ok=True)
    (output_root / "LATEST_RUN_PATH").write_text(str(sb), encoding="utf-8")

    required_scripts = [
        root / "build_rmbs_workbook.py",
        root / "rmbs_python_validation.py",
        root / "rmbs_v2_engine.py",
        root / "rmbs_excel_python_compare.py",
    ]
    ensure_files(required_scripts)

    env = dict(os.environ)
    env["PYTHONDONTWRITEBYTECODE"] = "1"

    preflight_lines = [
        f"timestamp={datetime.now().isoformat()}",
        f"python={sys.executable}",
        f"python_version={sys.version}",
        f"platform={platform.platform()}",
        f"pdflatex={shutil.which('pdflatex')}",
        f"output_root={output_root}",
        f"sandbox={sb}",
        f"automate_excel={args.automate_excel}",
        f"scenario_config={scenario_config_path if scenario_config_path else '(default)'}",
    ]
    (paths["logs"] / "preflight.log").write_text("\n".join(preflight_lines) + "\n", encoding="utf-8")

    results: list[CmdResult] = []
    build_cmd = [
        sys.executable,
        str(root / "build_rmbs_workbook.py"),
        "--output-path",
        str(sb / "rmbs_interview_model.xlsx"),
    ]
    if scenario_config_path:
        build_cmd.extend(["--scenario-config", str(scenario_config_path)])
    results.append(
        run_cmd(
            build_cmd,
            name="build_workbook",
            log_path=paths["logs"] / "build_workbook.log",
            env=env,
        )
    )
    v1_cmd = [
        sys.executable,
        str(root / "rmbs_python_validation.py"),
        "--out-report",
        str(paths["out_v1"] / "rmbs_validation_report.md"),
        "--out-summary-csv",
        str(paths["out_v1"] / "rmbs_validation_summary.csv"),
        "--out-monthly-csv",
        str(paths["out_v1"] / "rmbs_validation_monthly.csv"),
    ]
    if scenario_config_path:
        v1_cmd.extend(["--scenario-config", str(scenario_config_path)])
    results.append(
        run_cmd(
            v1_cmd,
            name="run_v1",
            log_path=paths["logs"] / "v1.log",
            env=env,
        )
    )
    v2_cmd = [
        sys.executable,
        str(root / "rmbs_v2_engine.py"),
        "--out-report",
        str(paths["out_v2"] / "rmbs_v2_validation_report.md"),
        "--out-summary-csv",
        str(paths["out_v2"] / "rmbs_v2_validation_summary.csv"),
        "--out-monthly-csv",
        str(paths["out_v2"] / "rmbs_v2_validation_monthly.csv"),
    ]
    if scenario_config_path:
        v2_cmd.extend(["--scenario-config", str(scenario_config_path)])
    results.append(
        run_cmd(
            v2_cmd,
            name="run_v2",
            log_path=paths["logs"] / "v2.log",
            env=env,
        )
    )

    for r in results:
        if r.returncode != 0:
            raise RuntimeError(f"Step failed: {r.name}. See log: {r.log_path}")

    workbook_path = sb / "rmbs_interview_model.xlsx"
    if args.automate_excel:
        excel_log = paths["logs"] / "excel_export_automation.log"
        with excel_log.open("w", encoding="utf-8") as log:
            for scenario in SCENARIOS:
                log.write(f"[excel] scenario={scenario} recalc+save\n")
                automate_excel_scenario_recalc(workbook_path, scenario)
                log.write(f"[export] scenario={scenario} csv\n")
                export_compare_csv_from_workbook(
                    workbook_path,
                    paths["inputs_exports"] / f"compare_export_{scenario}.csv",
                )

    compare_status_rows: list[list[object]] = []
    for scenario in SCENARIOS:
        export_csv = paths["inputs_exports"] / f"compare_export_{scenario}.csv"
        report = paths["out_compare"] / f"excel_python_{scenario}_report.md"
        diff = paths["out_compare"] / f"excel_python_{scenario}_diffs.csv"
        log = paths["logs"] / f"compare_{scenario}.log"
        if export_csv.exists():
            result = run_cmd(
                [
                    sys.executable,
                    str(root / "rmbs_excel_python_compare.py"),
                    "--excel-csv",
                    str(export_csv),
                    "--python-csv",
                    str(paths["out_v1"] / "rmbs_validation_monthly.csv"),
                    "--scenario",
                    scenario,
                    "--out-report",
                    str(report),
                    "--out-diff-csv",
                    str(diff),
                ],
                name=f"compare_{scenario}",
                log_path=log,
                env=env,
            )
            if result.returncode == 0:
                compare_status_rows.append([scenario, str(export_csv), "RAN", "ok"])
            else:
                compare_status_rows.append([scenario, str(export_csv), "FAILED", f"see {log}"])
        else:
            compare_status_rows.append(
                [scenario, str(export_csv), "MISSING", "manual Excel export required"]
            )
    write_csv(
        paths["logs"] / "compare_status.csv",
        ["scenario", "export_csv", "status", "detail"],
        compare_status_rows,
    )

    latex_rows: list[list[object]] = []
    latex_docs = [
        ("learning_report", root / "docs/latex/structured_finance_learning_report.tex", "latex_learning.log"),
        ("interview_qa", root / "docs/latex/interview_qa_deck.tex", "latex_qa.log"),
        ("atlas_summary", root / "docs/latex/atlas_summary.tex", "latex_atlas.log"),
        ("cheat_sheet", root / "docs/latex/cheat_sheet.tex", "latex_cheat.log"),
    ]
    if args.skip_latex:
        for name, _, log in latex_docs:
            latex_rows.append([name, "SKIPPED", str(paths["logs"] / log)])
    else:
        if shutil.which("pdflatex") is None:
            raise RuntimeError("pdflatex not found; install LaTeX or run with --skip-latex")
        for name, tex, log_name in latex_docs:
            log_path = paths["logs"] / log_name
            res = run_cmd(
                [
                    "pdflatex",
                    "-interaction=nonstopmode",
                    "-halt-on-error",
                    "-output-directory",
                    str(paths["out_docs"]),
                    str(tex),
                ],
                name=f"latex_{name}",
                log_path=log_path,
                env=env,
            )
            latex_rows.append([name, "OK" if res.returncode == 0 else "FAILED", str(log_path)])
    write_csv(paths["logs"] / "latex_status.csv", ["doc", "status", "log"], latex_rows)

    scorecard = build_scorecard(sb, strict=args.strict_tol, warn=args.warn_tol)
    score_json, score_md = write_scorecard_files(sb, scorecard)

    print(f"Sandbox: {sb}")
    print(f"Score: {scorecard['final_total']:.2f}/100")
    print(f"Grade: {scorecard['grade']}")
    print(f"Blocker: {'YES' if scorecard['blocker'] else 'NO'}")
    print(f"Scorecard MD: {score_md}")
    print(f"Scorecard JSON: {score_json}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
