#!/usr/bin/env python3
"""
Generate an entry-level RMBS Excel model with transparent formulas and checks.
"""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from scenario_config import load_scenario_bundle


def set_header(ws, row: int, headers: list[str]) -> None:
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = Font(bold=True)


def build_workbook(output_path: Path, scenario_bundle: dict | None = None) -> None:
    scenario_bundle = scenario_bundle or load_scenario_bundle(None)
    v1_scenarios = scenario_bundle["v1_scenarios"]
    v2_scenarios = scenario_bundle["v2_scenarios"]
    scenario_names = [str(item["name"]) for item in v1_scenarios]

    wb = Workbook()
    ws_inputs = wb.active
    ws_inputs.title = "Inputs"
    ws_pool = wb.create_sheet("LoanPool")
    ws_wf = wb.create_sheet("Waterfall")
    ws_scen = wb.create_sheet("Scenarios")
    ws_checks = wb.create_sheet("Checks")
    ws_scen_v2 = wb.create_sheet("Scenarios_V2")
    ws_pool_v2 = wb.create_sheet("PoolStates_V2")
    ws_wf_v2 = wb.create_sheet("Waterfall_V2")
    ws_checks_v2 = wb.create_sheet("Checks_V2")
    ws_export = wb.create_sheet("CompareExport")

    # ---------------- Scenarios ----------------
    set_header(ws_scen, 1, ["Scenario", "CPR_Annual", "CDR_Annual", "Severity"])
    for idx, scenario in enumerate(v1_scenarios, start=2):
        ws_scen[f"A{idx}"] = scenario["name"]
        ws_scen[f"B{idx}"] = float(scenario["cpr_annual"])
        ws_scen[f"C{idx}"] = float(scenario["cdr_annual"])
        ws_scen[f"D{idx}"] = float(scenario["severity"])

    ws_scen["A6"] = "Active scenario output"
    ws_scen["A6"].font = Font(bold=True)
    ws_scen["A7"] = "Scenario"
    ws_scen["B7"] = "=Inputs!B12"
    ws_scen["A8"] = "Cumulative Loss"
    ws_scen["B8"] = "=SUM(LoanPool!H2:H361)"
    ws_scen["A9"] = "Final Class A Balance"
    ws_scen["B9"] = "=Waterfall!X361"
    ws_scen["A10"] = "Final Class B Balance"
    ws_scen["B10"] = "=Waterfall!Y361"
    ws_scen["A11"] = "Final Equity Balance"
    ws_scen["B11"] = "=Waterfall!Z361"
    ws_scen["A12"] = "First Month Class A = 0"
    ws_scen["B12"] = '=IFERROR(MATCH(TRUE,Waterfall!X2:X361<=0.01,0),"Not Paid")'

    # ---------------- Scenarios_V2 ----------------
    set_header(
        ws_scen_v2,
        1,
        [
            "Scenario",
            "CPR_Annual",
            "Severity",
            "RollTo30_Annual",
            "DefaultFrom60_Annual",
            "Cure30_Monthly",
            "Roll30To60_Monthly",
            "Cure60_Monthly",
        ],
    )
    for idx, scenario in enumerate(v2_scenarios, start=2):
        ws_scen_v2[f"A{idx}"] = scenario["name"]
        ws_scen_v2[f"B{idx}"] = float(scenario["cpr_annual"])
        ws_scen_v2[f"C{idx}"] = float(scenario["severity"])
        ws_scen_v2[f"D{idx}"] = float(scenario["roll_to_30_annual"])
        ws_scen_v2[f"E{idx}"] = float(scenario["default_from_60_annual"])
        ws_scen_v2[f"F{idx}"] = float(scenario["cure_30_monthly"])
        ws_scen_v2[f"G{idx}"] = float(scenario["roll_30_to_60_monthly"])
        ws_scen_v2[f"H{idx}"] = float(scenario["cure_60_monthly"])

    ws_scen_v2["A6"] = "Active scenario output"
    ws_scen_v2["A6"].font = Font(bold=True)
    ws_scen_v2["A7"] = "Scenario"
    ws_scen_v2["B7"] = "=Inputs!B12"
    ws_scen_v2["A8"] = "Cumulative Loss (V2)"
    ws_scen_v2["B8"] = "=SUM(PoolStates_V2!Q2:Q361)"
    ws_scen_v2["A9"] = "Max DQ Ratio (of initial)"
    ws_scen_v2["B9"] = "=MAX(PoolStates_V2!Z2:Z361)"
    ws_scen_v2["A10"] = "First Trigger Month"
    ws_scen_v2["B10"] = '=IFERROR(MATCH(TRUE,PoolStates_V2!B2:B361=1,0),"Not Triggered")'
    ws_scen_v2["A11"] = "Final Class A Balance"
    ws_scen_v2["B11"] = "=Waterfall_V2!AP361"
    ws_scen_v2["A12"] = "Final Class B Balance"
    ws_scen_v2["B12"] = "=Waterfall_V2!AQ361"

    # ---------------- Inputs ----------------
    ws_inputs["A1"] = "RMBS Interview Model Inputs"
    ws_inputs["A1"].font = Font(bold=True)

    labels = [
        ("A2", "Initial Pool Balance"),
        ("A3", "Loan Count"),
        ("A4", "Balance Per Loan"),
        ("A5", "WAC (annual)"),
        ("A6", "Remaining Term (months)"),
        ("A7", "Servicing/Admin Fee (annual)"),
        ("A8", "Class A Share"),
        ("A9", "Class B Share"),
        ("A10", "Equity Share"),
        ("A12", "Active Scenario"),
        ("A13", "CPR (annual, active scenario)"),
        ("A14", "CDR (annual, active scenario)"),
        ("A15", "Severity (active scenario)"),
        ("A16", "SMM (monthly)"),
        ("A17", "MDR (monthly)"),
        ("A23", "V2 Trigger Loss Threshold"),
        ("A24", "V2 Trigger DQ Threshold"),
        ("A25", "V2 Initial DQ30 Share"),
        ("A26", "V2 Initial DQ60 Share"),
        ("A27", "V2 Roll to 30 (annual)"),
        ("A28", "V2 Default from 60 (annual)"),
        ("A29", "V2 Cure30 (monthly)"),
        ("A30", "V2 Roll30->60 (monthly)"),
        ("A31", "V2 Cure60 (monthly)"),
        ("A32", "V2 SMM (monthly)"),
        ("A33", "V2 RollTo30 MDR (monthly)"),
        ("A34", "V2 Default60 MDR (monthly)"),
        ("A35", "V2 Severity (active scenario)"),
    ]
    for cell, text in labels:
        ws_inputs[cell] = text

    ws_inputs["B3"] = 1000
    ws_inputs["B4"] = 300000
    ws_inputs["B2"] = "=B3*B4"
    ws_inputs["B5"] = 0.0575
    ws_inputs["B6"] = 300
    ws_inputs["B7"] = 0.005
    ws_inputs["B8"] = 0.8
    ws_inputs["B9"] = 0.15
    ws_inputs["B10"] = 0.05
    ws_inputs["B12"] = scenario_names[0]
    ws_inputs["B13"] = '=INDEX(Scenarios!B2:B4,MATCH(B12,Scenarios!A2:A4,0))'
    ws_inputs["B14"] = '=INDEX(Scenarios!C2:C4,MATCH(B12,Scenarios!A2:A4,0))'
    ws_inputs["B15"] = '=INDEX(Scenarios!D2:D4,MATCH(B12,Scenarios!A2:A4,0))'
    ws_inputs["B16"] = "=1-(1-B13)^(1/12)"
    ws_inputs["B17"] = "=1-(1-B14)^(1/12)"
    ws_inputs["B23"] = 0.03
    ws_inputs["B24"] = 0.06
    ws_inputs["B25"] = 0.01
    ws_inputs["B26"] = 0.005
    ws_inputs["B27"] = '=INDEX(Scenarios_V2!D2:D4,MATCH(B12,Scenarios_V2!A2:A4,0))'
    ws_inputs["B28"] = '=INDEX(Scenarios_V2!E2:E4,MATCH(B12,Scenarios_V2!A2:A4,0))'
    ws_inputs["B29"] = '=INDEX(Scenarios_V2!F2:F4,MATCH(B12,Scenarios_V2!A2:A4,0))'
    ws_inputs["B30"] = '=INDEX(Scenarios_V2!G2:G4,MATCH(B12,Scenarios_V2!A2:A4,0))'
    ws_inputs["B31"] = '=INDEX(Scenarios_V2!H2:H4,MATCH(B12,Scenarios_V2!A2:A4,0))'
    ws_inputs["B32"] = '=INDEX(Scenarios_V2!B2:B4,MATCH(B12,Scenarios_V2!A2:A4,0))'
    ws_inputs["B33"] = "=1-(1-B27)^(1/12)"
    ws_inputs["B34"] = "=1-(1-B28)^(1/12)"
    ws_inputs["B35"] = '=INDEX(Scenarios_V2!C2:C4,MATCH(B12,Scenarios_V2!A2:A4,0))'

    ws_inputs["A37"] = "Notes"
    ws_inputs["A37"].font = Font(bold=True)
    ws_inputs["A38"] = "All downstream formulas should reference these inputs."
    ws_inputs["A39"] = "Change B12 to Base, Mild_Stress, or Severe_Stress."

    dv = DataValidation(type="list", formula1=f'"{",".join(scenario_names)}"', allow_blank=False)
    ws_inputs.add_data_validation(dv)
    dv.add(ws_inputs["B12"])

    # ---------------- LoanPool ----------------
    pool_headers = [
        "Month",
        "Begin_Balance",
        "Scheduled_Payment",
        "Interest",
        "Scheduled_Principal",
        "Prepayment",
        "Default",
        "Loss",
        "Total_Principal_Outflow",
        "End_Balance",
        "Fee",
        "Interest_Available",
        "Principal_Available",
    ]
    set_header(ws_pool, 1, pool_headers)

    for r in range(2, 362):
        month = r - 1
        ws_pool.cell(row=r, column=1, value=month)
        if r == 2:
            ws_pool.cell(row=r, column=2, value="=Inputs!B2")
        else:
            ws_pool.cell(row=r, column=2, value=f"=J{r-1}")

        ws_pool.cell(
            row=r,
            column=3,
            value=f"=IF(A{r}<=Inputs!B6,IF(B{r}>0,-PMT(Inputs!B5/12,Inputs!B6-A{r}+1,B{r}),0),0)",
        )
        ws_pool.cell(row=r, column=4, value=f"=B{r}*Inputs!B5/12")
        ws_pool.cell(row=r, column=5, value=f"=MAX(C{r}-D{r},0)")
        ws_pool.cell(row=r, column=6, value=f"=MAX((B{r}-E{r})*Inputs!B16,0)")
        ws_pool.cell(row=r, column=7, value=f"=MAX((B{r}-E{r}-F{r})*Inputs!B17,0)")
        ws_pool.cell(row=r, column=8, value=f"=G{r}*Inputs!B15")
        ws_pool.cell(row=r, column=9, value=f"=MIN(B{r},E{r}+F{r}+G{r})")
        ws_pool.cell(row=r, column=10, value=f"=MAX(B{r}-I{r},0)")
        ws_pool.cell(row=r, column=11, value=f"=B{r}*Inputs!B7/12")
        ws_pool.cell(row=r, column=12, value=f"=MAX(D{r}-K{r},0)")
        ws_pool.cell(row=r, column=13, value=f"=MAX(I{r}-H{r},0)")

    # ---------------- Waterfall ----------------
    wf_headers = [
        "Month",
        "Pool_Interest_Available",
        "Pool_Principal_Available",
        "Pool_Loss",
        "A_Begin",
        "B_Begin",
        "E_Begin",
        "Loss_to_E",
        "Loss_to_B",
        "Loss_to_A",
        "A_After_Loss",
        "B_After_Loss",
        "E_After_Loss",
        "A_Int_Due",
        "A_Int_Paid",
        "A_Int_Shortfall",
        "B_Int_Due",
        "B_Int_Paid",
        "B_Int_Shortfall",
        "E_Residual_Int",
        "A_Principal_Paid",
        "B_Principal_Paid",
        "E_Principal_Paid",
        "A_End",
        "B_End",
        "E_End",
        "Principal_Diff",
        "Interest_Diff",
        "Loss_Diff",
    ]
    set_header(ws_wf, 1, wf_headers)

    for r in range(2, 362):
        ws_wf.cell(row=r, column=1, value=r - 1)
        ws_wf.cell(row=r, column=2, value=f"=LoanPool!L{r}")
        ws_wf.cell(row=r, column=3, value=f"=LoanPool!M{r}")
        ws_wf.cell(row=r, column=4, value=f"=LoanPool!H{r}")

        if r == 2:
            ws_wf.cell(row=r, column=5, value="=Inputs!B2*Inputs!B8")
            ws_wf.cell(row=r, column=6, value="=Inputs!B2*Inputs!B9")
            ws_wf.cell(row=r, column=7, value="=Inputs!B2*Inputs!B10")
        else:
            ws_wf.cell(row=r, column=5, value=f"=X{r-1}")
            ws_wf.cell(row=r, column=6, value=f"=Y{r-1}")
            ws_wf.cell(row=r, column=7, value=f"=Z{r-1}")

        ws_wf.cell(row=r, column=8, value=f"=MIN(D{r},G{r})")
        ws_wf.cell(row=r, column=9, value=f"=MIN(MAX(D{r}-H{r},0),F{r})")
        ws_wf.cell(row=r, column=10, value=f"=MIN(MAX(D{r}-H{r}-I{r},0),E{r})")
        ws_wf.cell(row=r, column=11, value=f"=MAX(E{r}-J{r},0)")
        ws_wf.cell(row=r, column=12, value=f"=MAX(F{r}-I{r},0)")
        ws_wf.cell(row=r, column=13, value=f"=MAX(G{r}-H{r},0)")
        ws_wf.cell(row=r, column=14, value=f"=K{r}*Inputs!B5/12")
        ws_wf.cell(row=r, column=15, value=f"=MIN(B{r},N{r})")
        ws_wf.cell(row=r, column=16, value=f"=N{r}-O{r}")
        ws_wf.cell(row=r, column=17, value=f"=L{r}*Inputs!B5/12")
        ws_wf.cell(row=r, column=18, value=f"=MIN(MAX(B{r}-O{r},0),Q{r})")
        ws_wf.cell(row=r, column=19, value=f"=Q{r}-R{r}")
        ws_wf.cell(row=r, column=20, value=f"=MAX(B{r}-O{r}-R{r},0)")
        ws_wf.cell(row=r, column=21, value=f"=MIN(C{r},K{r})")
        ws_wf.cell(row=r, column=22, value=f"=MIN(MAX(C{r}-U{r},0),L{r})")
        ws_wf.cell(row=r, column=23, value=f"=MIN(MAX(C{r}-U{r}-V{r},0),M{r})")
        ws_wf.cell(row=r, column=24, value=f"=MAX(K{r}-U{r},0)")
        ws_wf.cell(row=r, column=25, value=f"=MAX(L{r}-V{r},0)")
        ws_wf.cell(row=r, column=26, value=f"=MAX(M{r}-W{r},0)")
        ws_wf.cell(row=r, column=27, value=f"=C{r}-(U{r}+V{r}+W{r})")
        ws_wf.cell(row=r, column=28, value=f"=B{r}-(O{r}+R{r}+T{r})")
        ws_wf.cell(row=r, column=29, value=f"=D{r}-(H{r}+I{r}+J{r})")

    # ---------------- Checks ----------------
    set_header(ws_checks, 1, ["Check", "Metric", "Status", "Target"])

    ws_checks["A2"] = "Principal allocation closes"
    ws_checks["B2"] = "=MAX(MAX(Waterfall!AA2:AA361),-MIN(Waterfall!AA2:AA361))"
    ws_checks["C2"] = '=IF(B2<=0.01,"PASS","FAIL")'
    ws_checks["D2"] = "<= 0.01"

    ws_checks["A3"] = "Interest allocation closes"
    ws_checks["B3"] = "=MAX(MAX(Waterfall!AB2:AB361),-MIN(Waterfall!AB2:AB361))"
    ws_checks["C3"] = '=IF(B3<=0.01,"PASS","FAIL")'
    ws_checks["D3"] = "<= 0.01"

    ws_checks["A4"] = "Loss allocation closes"
    ws_checks["B4"] = "=MAX(MAX(Waterfall!AC2:AC361),-MIN(Waterfall!AC2:AC361))"
    ws_checks["C4"] = '=IF(B4<=0.01,"PASS","FAIL")'
    ws_checks["D4"] = "<= 0.01"

    ws_checks["A5"] = "No negative balances"
    ws_checks["B5"] = "=MIN(LoanPool!J2:J361,Waterfall!X2:X361,Waterfall!Y2:Y361,Waterfall!Z2:Z361)"
    ws_checks["C5"] = '=IF(B5>=-0.01,"PASS","FAIL")'
    ws_checks["D5"] = ">= -0.01"

    ws_checks["A6"] = "Scenario ordering (stress increases credit loss params)"
    ws_checks["B6"] = '=IF(AND(Scenarios!B2>=Scenarios!B3,Scenarios!B3>=Scenarios!B4,Scenarios!C2<=Scenarios!C3,Scenarios!C3<=Scenarios!C4,Scenarios!D2<=Scenarios!D3,Scenarios!D3<=Scenarios!D4),1,0)'
    ws_checks["C6"] = '=IF(B6=1,"PASS","FAIL")'
    ws_checks["D6"] = "1"

    ws_checks["A8"] = "How to use"
    ws_checks["A8"].font = Font(bold=True)
    ws_checks["A9"] = "1) Choose scenario in Inputs!B12."
    ws_checks["A10"] = "2) Recalculate workbook."
    ws_checks["A11"] = "3) Review this tab and Scenarios tab outputs."

    # ---------------- PoolStates_V2 ----------------
    pool_v2_headers = [
        "Month",
        "Trigger_On",
        "Pool_Begin",
        "Current_Begin",
        "DQ30_Begin",
        "DQ60_Begin",
        "Remaining_Term",
        "Sched_Payment",
        "Interest",
        "Sched_Principal",
        "Prepay",
        "Roll_to_30",
        "Cure30",
        "Roll30_to_60",
        "Cure60",
        "Default",
        "Loss",
        "Fee",
        "Interest_Available",
        "Principal_From_Pool",
        "Current_End",
        "DQ30_End",
        "DQ60_End",
        "Pool_End",
        "Cum_Loss",
        "DQ_Ratio_Init",
        "Trigger_Event",
        "Pool_Roll_Diff",
    ]
    set_header(ws_pool_v2, 1, pool_v2_headers)
    for r in range(2, 362):
        ws_pool_v2.cell(row=r, column=1, value=r - 1)
        if r == 2:
            ws_pool_v2.cell(row=r, column=2, value=f"=AA{r}")
            ws_pool_v2.cell(row=r, column=3, value="=Inputs!B2")
            ws_pool_v2.cell(row=r, column=4, value="=Inputs!B2*(1-Inputs!B25-Inputs!B26)")
            ws_pool_v2.cell(row=r, column=5, value="=Inputs!B2*Inputs!B25")
            ws_pool_v2.cell(row=r, column=6, value="=Inputs!B2*Inputs!B26")
            ws_pool_v2.cell(row=r, column=25, value=f"=Q{r}")
        else:
            ws_pool_v2.cell(row=r, column=2, value=f"=MAX(B{r-1},AA{r})")
            ws_pool_v2.cell(row=r, column=3, value=f"=X{r-1}")
            ws_pool_v2.cell(row=r, column=4, value=f"=U{r-1}")
            ws_pool_v2.cell(row=r, column=5, value=f"=V{r-1}")
            ws_pool_v2.cell(row=r, column=6, value=f"=W{r-1}")
            ws_pool_v2.cell(row=r, column=25, value=f"=Y{r-1}+Q{r}")

        ws_pool_v2.cell(row=r, column=7, value=f"=MAX(Inputs!B6-A{r}+1,0)")
        ws_pool_v2.cell(
            row=r,
            column=8,
            value=f"=IF(AND(A{r}<=Inputs!B6,D{r}>0),-PMT(Inputs!B5/12,G{r},D{r}),0)",
        )
        ws_pool_v2.cell(row=r, column=9, value=f"=IF(AND(A{r}<=Inputs!B6,D{r}>0),D{r}*Inputs!B5/12,0)")
        ws_pool_v2.cell(row=r, column=10, value=f"=MAX(H{r}-I{r},0)")
        ws_pool_v2.cell(row=r, column=11, value=f"=MAX((D{r}-J{r})*Inputs!B32,0)")
        ws_pool_v2.cell(row=r, column=12, value=f"=MAX((D{r}-J{r}-K{r})*Inputs!B33,0)")
        ws_pool_v2.cell(row=r, column=13, value=f"=MIN(E{r}*Inputs!B29,E{r})")
        ws_pool_v2.cell(row=r, column=14, value=f"=MIN((E{r}-M{r})*Inputs!B30,E{r}-M{r})")
        ws_pool_v2.cell(row=r, column=15, value=f"=MIN(F{r}*Inputs!B31,F{r})")
        ws_pool_v2.cell(row=r, column=16, value=f"=MIN((F{r}-O{r})*Inputs!B34,F{r}-O{r})")
        ws_pool_v2.cell(row=r, column=17, value=f"=P{r}*Inputs!B35")
        ws_pool_v2.cell(row=r, column=18, value=f"=C{r}*Inputs!B7/12")
        ws_pool_v2.cell(row=r, column=19, value=f"=MAX(I{r}-R{r},0)")
        ws_pool_v2.cell(row=r, column=20, value=f"=MAX(J{r}+K{r}+P{r}-Q{r},0)")
        ws_pool_v2.cell(row=r, column=21, value=f"=MAX(D{r}-J{r}-K{r}-L{r}+M{r}+O{r},0)")
        ws_pool_v2.cell(row=r, column=22, value=f"=MAX(E{r}-M{r}-N{r}+L{r},0)")
        ws_pool_v2.cell(row=r, column=23, value=f"=MAX(F{r}-O{r}-P{r}+N{r},0)")
        ws_pool_v2.cell(row=r, column=24, value=f"=U{r}+V{r}+W{r}")
        ws_pool_v2.cell(row=r, column=26, value=f"=IF(Inputs!B2>0,(V{r}+W{r})/Inputs!B2,0)")
        ws_pool_v2.cell(
            row=r,
            column=27,
            value=f"=IF(OR(Y{r}/Inputs!B2>=Inputs!B23,Z{r}>=Inputs!B24),1,0)",
        )
        ws_pool_v2.cell(row=r, column=28, value=f"=MAX(C{r}-J{r}-K{r}-P{r},0)-X{r}")

    # ---------------- Waterfall_V2 ----------------
    wf_v2_headers = [
        "Month",
        "Trigger_On",
        "Interest_Available",
        "Principal_From_Pool",
        "Loss",
        "A_Begin",
        "B_Begin",
        "E_Begin",
        "Loss_to_E",
        "Loss_to_B",
        "Loss_to_A",
        "Unalloc_Loss",
        "A_After_Loss",
        "B_After_Loss",
        "E_After_Loss",
        "A_Int_Due",
        "A_Int_Paid",
        "B_Int_Due",
        "B_Int_Paid",
        "Residual_Interest",
        "Trapped_Interest",
        "E_Residual_Int",
        "Principal_Available",
        "AB_Total",
        "A_Target_ProRata",
        "B_Target_ProRata",
        "A_Prin_ProRata",
        "B_Prin_ProRata",
        "Rem_After_ProRata",
        "Extra_A",
        "Rem_After_ExtraA",
        "Extra_B",
        "Rem_After_ExtraB",
        "E_Prin_ProRata",
        "A_Prin_Seq",
        "B_Prin_Seq",
        "E_Prin_Seq",
        "A_Principal_Paid",
        "B_Principal_Paid",
        "E_Principal_Paid",
        "Unalloc_Principal",
        "A_End",
        "B_End",
        "E_End",
        "Principal_Diff",
        "Interest_Diff",
        "Loss_Diff",
    ]
    set_header(ws_wf_v2, 1, wf_v2_headers)
    for r in range(2, 362):
        ws_wf_v2.cell(row=r, column=1, value=r - 1)
        ws_wf_v2.cell(row=r, column=2, value=f"=PoolStates_V2!B{r}")
        ws_wf_v2.cell(row=r, column=3, value=f"=PoolStates_V2!S{r}")
        ws_wf_v2.cell(row=r, column=4, value=f"=PoolStates_V2!T{r}")
        ws_wf_v2.cell(row=r, column=5, value=f"=PoolStates_V2!Q{r}")

        if r == 2:
            ws_wf_v2.cell(row=r, column=6, value="=Inputs!B2*Inputs!B8")
            ws_wf_v2.cell(row=r, column=7, value="=Inputs!B2*Inputs!B9")
            ws_wf_v2.cell(row=r, column=8, value="=Inputs!B2*Inputs!B10")
        else:
            ws_wf_v2.cell(row=r, column=6, value=f"=AP{r-1}")
            ws_wf_v2.cell(row=r, column=7, value=f"=AQ{r-1}")
            ws_wf_v2.cell(row=r, column=8, value=f"=AR{r-1}")

        ws_wf_v2.cell(row=r, column=9, value=f"=MIN(E{r},H{r})")
        ws_wf_v2.cell(row=r, column=10, value=f"=MIN(MAX(E{r}-I{r},0),G{r})")
        ws_wf_v2.cell(row=r, column=11, value=f"=MIN(MAX(E{r}-I{r}-J{r},0),F{r})")
        ws_wf_v2.cell(row=r, column=12, value=f"=MAX(E{r}-I{r}-J{r}-K{r},0)")
        ws_wf_v2.cell(row=r, column=13, value=f"=MAX(F{r}-K{r},0)")
        ws_wf_v2.cell(row=r, column=14, value=f"=MAX(G{r}-J{r},0)")
        ws_wf_v2.cell(row=r, column=15, value=f"=MAX(H{r}-I{r},0)")
        ws_wf_v2.cell(row=r, column=16, value=f"=M{r}*Inputs!B5/12")
        ws_wf_v2.cell(row=r, column=17, value=f"=MIN(C{r},P{r})")
        ws_wf_v2.cell(row=r, column=18, value=f"=N{r}*Inputs!B5/12")
        ws_wf_v2.cell(row=r, column=19, value=f"=MIN(MAX(C{r}-Q{r},0),R{r})")
        ws_wf_v2.cell(row=r, column=20, value=f"=MAX(C{r}-Q{r}-S{r},0)")
        ws_wf_v2.cell(row=r, column=21, value=f"=IF(B{r}=1,T{r},0)")
        ws_wf_v2.cell(row=r, column=22, value=f"=T{r}-U{r}")
        ws_wf_v2.cell(row=r, column=23, value=f"=D{r}+U{r}")
        ws_wf_v2.cell(row=r, column=24, value=f"=M{r}+N{r}")
        ws_wf_v2.cell(row=r, column=25, value=f"=IF(X{r}>0,W{r}*(M{r}/X{r}),0)")
        ws_wf_v2.cell(row=r, column=26, value=f"=IF(X{r}>0,W{r}-Y{r},0)")
        ws_wf_v2.cell(row=r, column=27, value=f"=MIN(Y{r},M{r})")
        ws_wf_v2.cell(row=r, column=28, value=f"=MIN(Z{r},N{r})")
        ws_wf_v2.cell(row=r, column=29, value=f"=MAX(W{r}-AA{r}-AB{r},0)")
        ws_wf_v2.cell(row=r, column=30, value=f"=MIN(AC{r},MAX(M{r}-AA{r},0))")
        ws_wf_v2.cell(row=r, column=31, value=f"=MAX(AC{r}-AD{r},0)")
        ws_wf_v2.cell(row=r, column=32, value=f"=MIN(AE{r},MAX(N{r}-AB{r},0))")
        ws_wf_v2.cell(row=r, column=33, value=f"=MAX(AE{r}-AF{r},0)")
        ws_wf_v2.cell(row=r, column=34, value=f"=MIN(AG{r},O{r})")
        ws_wf_v2.cell(row=r, column=35, value=f"=MIN(W{r},M{r})")
        ws_wf_v2.cell(row=r, column=36, value=f"=MIN(MAX(W{r}-AI{r},0),N{r})")
        ws_wf_v2.cell(row=r, column=37, value=f"=MIN(MAX(W{r}-AI{r}-AJ{r},0),O{r})")
        ws_wf_v2.cell(row=r, column=38, value=f"=IF(B{r}=1,AI{r},AA{r}+AD{r})")
        ws_wf_v2.cell(row=r, column=39, value=f"=IF(B{r}=1,AJ{r},AB{r}+AF{r})")
        ws_wf_v2.cell(row=r, column=40, value=f"=IF(B{r}=1,AK{r},AH{r})")
        ws_wf_v2.cell(row=r, column=41, value=f"=MAX(W{r}-AL{r}-AM{r}-AN{r},0)")
        ws_wf_v2.cell(row=r, column=42, value=f"=MAX(M{r}-AL{r},0)")
        ws_wf_v2.cell(row=r, column=43, value=f"=MAX(N{r}-AM{r},0)")
        ws_wf_v2.cell(row=r, column=44, value=f"=MAX(O{r}-AN{r},0)")
        ws_wf_v2.cell(row=r, column=45, value=f"=W{r}-(AL{r}+AM{r}+AN{r}+AO{r})")
        ws_wf_v2.cell(row=r, column=46, value=f"=C{r}-(Q{r}+S{r}+V{r}+U{r})")
        ws_wf_v2.cell(row=r, column=47, value=f"=E{r}-(I{r}+J{r}+K{r}+L{r})")

    # ---------------- Checks_V2 ----------------
    set_header(ws_checks_v2, 1, ["Check", "Metric", "Status", "Target"])
    ws_checks_v2["A2"] = "Principal allocation closes (V2)"
    ws_checks_v2["B2"] = "=MAX(MAX(Waterfall_V2!AS2:AS361),-MIN(Waterfall_V2!AS2:AS361))"
    ws_checks_v2["C2"] = '=IF(B2<=0.01,"PASS","FAIL")'
    ws_checks_v2["D2"] = "<= 0.01"

    ws_checks_v2["A3"] = "Interest allocation closes (V2)"
    ws_checks_v2["B3"] = "=MAX(MAX(Waterfall_V2!AT2:AT361),-MIN(Waterfall_V2!AT2:AT361))"
    ws_checks_v2["C3"] = '=IF(B3<=0.01,"PASS","FAIL")'
    ws_checks_v2["D3"] = "<= 0.01"

    ws_checks_v2["A4"] = "Loss allocation closes (V2)"
    ws_checks_v2["B4"] = "=MAX(MAX(Waterfall_V2!AU2:AU361),-MIN(Waterfall_V2!AU2:AU361))"
    ws_checks_v2["C4"] = '=IF(B4<=0.01,"PASS","FAIL")'
    ws_checks_v2["D4"] = "<= 0.01"

    ws_checks_v2["A5"] = "Pool roll-forward closes (V2)"
    ws_checks_v2["B5"] = "=MAX(MAX(PoolStates_V2!AB2:AB361),-MIN(PoolStates_V2!AB2:AB361))"
    ws_checks_v2["C5"] = '=IF(B5<=0.01,"PASS","FAIL")'
    ws_checks_v2["D5"] = "<= 0.01"

    ws_checks_v2["A6"] = "No negative balances (V2)"
    ws_checks_v2["B6"] = "=MIN(PoolStates_V2!U2:U361,PoolStates_V2!V2:V361,PoolStates_V2!W2:W361,Waterfall_V2!AP2:AP361,Waterfall_V2!AQ2:AQ361,Waterfall_V2!AR2:AR361)"
    ws_checks_v2["C6"] = '=IF(B6>=-0.01,"PASS","FAIL")'
    ws_checks_v2["D6"] = ">= -0.01"

    ws_checks_v2["A7"] = "Trigger persistence (no 1->0)"
    ws_checks_v2["B7"] = "=SUMPRODUCT(--(PoolStates_V2!B3:B361<PoolStates_V2!B2:B360))"
    ws_checks_v2["C7"] = '=IF(B7=0,"PASS","FAIL")'
    ws_checks_v2["D7"] = "0"

    ws_checks_v2["A8"] = "Trigger event consistency"
    ws_checks_v2["B8"] = "=SUMPRODUCT(--(PoolStates_V2!B2:B361<PoolStates_V2!AA2:AA361))"
    ws_checks_v2["C8"] = '=IF(B8=0,"PASS","FAIL")'
    ws_checks_v2["D8"] = "0"

    ws_checks_v2["A9"] = "Stress parameter ordering (V2)"
    ws_checks_v2["B9"] = '=IF(AND(Scenarios_V2!B2>=Scenarios_V2!B3,Scenarios_V2!B3>=Scenarios_V2!B4,Scenarios_V2!C2<=Scenarios_V2!C3,Scenarios_V2!C3<=Scenarios_V2!C4,Scenarios_V2!D2<=Scenarios_V2!D3,Scenarios_V2!D3<=Scenarios_V2!D4,Scenarios_V2!E2<=Scenarios_V2!E3,Scenarios_V2!E3<=Scenarios_V2!E4),1,0)'
    ws_checks_v2["C9"] = '=IF(B9=1,"PASS","FAIL")'
    ws_checks_v2["D9"] = "1"

    ws_checks_v2["A11"] = "How to use (V2)"
    ws_checks_v2["A11"].font = Font(bold=True)
    ws_checks_v2["A12"] = "1) Select scenario in Inputs!B12."
    ws_checks_v2["A13"] = "2) Recalculate workbook."
    ws_checks_v2["A14"] = "3) Review Checks_V2 and Scenarios_V2 outputs."

    # ---------------- CompareExport ----------------
    export_headers = [
        "scenario",
        "month",
        "pool_begin",
        "pool_end",
        "interest_available",
        "principal_available",
        "loss",
        "a_end",
        "b_end",
        "e_end",
        "principal_diff",
        "interest_diff",
        "loss_diff",
    ]
    set_header(ws_export, 1, export_headers)
    for r in range(2, 362):
        ws_export.cell(row=r, column=1, value="=Inputs!B12")
        ws_export.cell(row=r, column=2, value=f"=LoanPool!A{r}")
        ws_export.cell(row=r, column=3, value=f"=LoanPool!B{r}")
        ws_export.cell(row=r, column=4, value=f"=LoanPool!J{r}")
        ws_export.cell(row=r, column=5, value=f"=LoanPool!L{r}")
        ws_export.cell(row=r, column=6, value=f"=LoanPool!M{r}")
        ws_export.cell(row=r, column=7, value=f"=LoanPool!H{r}")
        ws_export.cell(row=r, column=8, value=f"=Waterfall!X{r}")
        ws_export.cell(row=r, column=9, value=f"=Waterfall!Y{r}")
        ws_export.cell(row=r, column=10, value=f"=Waterfall!Z{r}")
        ws_export.cell(row=r, column=11, value=f"=Waterfall!AA{r}")
        ws_export.cell(row=r, column=12, value=f"=Waterfall!AB{r}")
        ws_export.cell(row=r, column=13, value=f"=Waterfall!AC{r}")

    ws_export["A364"] = "Export Instructions"
    ws_export["A364"].font = Font(bold=True)
    ws_export["A365"] = "1) Recalculate workbook in Excel."
    ws_export["A366"] = "2) Keep one active scenario in Inputs!B12."
    ws_export["A367"] = "3) Save CompareExport sheet as CSV."
    ws_export["A368"] = "4) Run rmbs_excel_python_compare.py with that CSV."

    # Formatting
    for ws in [ws_inputs, ws_pool, ws_wf, ws_scen, ws_checks, ws_scen_v2, ws_pool_v2, ws_wf_v2, ws_checks_v2, ws_export]:
        ws.freeze_panes = "A2"
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

    # Better widths for wide sheets
    ws_wf.column_dimensions["A"].width = 10
    for col in "BCDEFGHIJKLMNOPQRSTUVWXYZ":
        ws_wf.column_dimensions[col].width = 16
    ws_pool.column_dimensions["A"].width = 10
    for col in "BCDEFGHIJKLM":
        ws_pool.column_dimensions[col].width = 18
    ws_pool_v2.column_dimensions["A"].width = 10
    for col in [
        "B",
        "C",
        "D",
        "E",
        "F",
        "G",
        "H",
        "I",
        "J",
        "K",
        "L",
        "M",
        "N",
        "O",
        "P",
        "Q",
        "R",
        "S",
        "T",
        "U",
        "V",
        "W",
        "X",
        "Y",
        "Z",
        "AA",
        "AB",
    ]:
        ws_pool_v2.column_dimensions[col].width = 15
    ws_wf_v2.column_dimensions["A"].width = 10
    for col_idx in range(2, ws_wf_v2.max_column + 1):
        ws_wf_v2.column_dimensions[get_column_letter(col_idx)].width = 14
    ws_export.column_dimensions["A"].width = 18
    ws_export.column_dimensions["B"].width = 10
    for col in "CDEFGHIJKLM":
        ws_export.column_dimensions[col].width = 17

    wb.save(output_path)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Build RMBS workbook.")
    parser.add_argument(
        "--output-path",
        default="rmbs_waterfall_model.xlsx",
        help="Output workbook path.",
    )
    parser.add_argument(
        "--scenario-config",
        default="",
        help="Optional YAML path with v1_scenarios/v2_scenarios. If omitted, defaults are used.",
    )
    args = parser.parse_args()

    out = Path(args.output_path)
    bundle = load_scenario_bundle(args.scenario_config or None)
    build_workbook(out, bundle)
    print(f"Wrote {out.resolve()}")
